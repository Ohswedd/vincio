"""Real-behavior coverage tests for ``vincio.documents.loaders``.

Exercises the loader dispatch, every dependency-free built-in format, the
optional-dependency error paths (pypdf / openpyxl absent in the offline test
env), the audio/figure ingestion paths, and the directory walker — all offline
and deterministic. No mock/patch: real files on disk, real parsers, real
``MockTranscriber`` / ``MockVideoAnalyzer`` and small concrete stub engines.
"""

from __future__ import annotations

from pathlib import Path

import pytest

from vincio.core.errors import DocumentError, LoaderError
from vincio.core.types import Document
from vincio.documents.audio import MockTranscriber
from vincio.documents.loaders import (
    _default_rasterizer,
    _read_text,
    figure_evidence,
    load_directory,
    load_document,
    load_docx,
    load_media,
    load_pdf,
    load_video,
    supported_extensions,
)
from vincio.documents.multimodal import ImageObservation
from vincio.documents.video import MockVideoAnalyzer

# whether the optional heavy parsers are importable in this environment
try:  # pragma: no cover - environment probe
    import openpyxl  # noqa: F401

    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    _HAS_OPENPYXL = False


# -- _read_text encoding fallback ---------------------------------------------


def test_read_text_falls_back_to_latin1(tmp_path):
    p = tmp_path / "weird.txt"
    # 0xE9 ('é' in latin-1) is not valid standalone UTF-8 → triggers the fallback.
    p.write_bytes(b"caf\xe9 noir")
    assert _read_text(p) == "café noir"


def test_load_document_latin1_text_file(tmp_path):
    p = tmp_path / "n.txt"
    p.write_bytes(b"r\xe9sum\xe9")
    doc = load_document(p)
    assert doc.text == "résumé"
    assert doc.media_type == "text/plain"


# -- plain text / markdown ----------------------------------------------------


def test_txt_is_plain_no_sections(tmp_path):
    p = tmp_path / "note.txt"
    p.write_text("just a flat note", encoding="utf-8")
    doc = load_document(p)
    assert doc.media_type == "text/plain"
    assert doc.sections == [] and doc.tables == []
    assert doc.title == "note"


def test_markdown_extracts_sections_and_tables(tmp_path):
    p = tmp_path / "doc.md"
    p.write_text(
        "# Heading One\n\nbody text\n\n"
        "| A | B |\n|---|---|\n| 1 | 2 |\n",
        encoding="utf-8",
    )
    doc = load_document(p)
    assert doc.media_type == "text/markdown"
    titles = [s["title"] for s in doc.sections]
    assert "Heading One" in titles
    assert len(doc.tables) == 1
    assert doc.tables[0]["columns"] == ["A", "B"]


def test_rst_treated_as_markdown(tmp_path):
    p = tmp_path / "r.rst"
    p.write_text("# Title\n\ncontent\n", encoding="utf-8")
    doc = load_document(p)
    assert doc.media_type == "text/markdown"


# -- HTML ---------------------------------------------------------------------


def test_html_extracts_title_text_and_table(tmp_path):
    p = tmp_path / "page.html"
    p.write_text(
        "<html><head><title>My Page</title></head><body>"
        "<h1>Header</h1><p>Paragraph body.</p>"
        "<table><tr><th>Name</th><th>Qty</th></tr>"
        "<tr><td>Apple</td><td>3</td></tr></table>"
        "</body></html>",
        encoding="utf-8",
    )
    doc = load_document(p)
    assert doc.media_type == "text/html"
    assert doc.title == "My Page"
    assert "Paragraph body." in doc.text
    assert len(doc.tables) == 1
    assert doc.tables[0]["columns"] == ["Name", "Qty"]


def test_html_falls_back_to_stem_when_no_title(tmp_path):
    p = tmp_path / "untitled.htm"
    p.write_text("<body><p>only body</p></body>", encoding="utf-8")
    doc = load_document(p)
    assert doc.title == "untitled"


# -- CSV / TSV ----------------------------------------------------------------


def test_csv_parses_into_table(tmp_path):
    p = tmp_path / "data.csv"
    p.write_text("name,age\nAda,36\nGrace,40\n", encoding="utf-8")
    doc = load_document(p)
    assert doc.media_type == "text/csv"
    table = doc.tables[0]
    assert table["columns"] == ["name", "age"]
    assert table["rows"] == [["Ada", "36"], ["Grace", "40"]]


def test_tsv_uses_tab_delimiter(tmp_path):
    p = tmp_path / "data.tsv"
    p.write_text("a\tb\n1\t2\n", encoding="utf-8")
    doc = load_document(p)
    assert doc.tables[0]["columns"] == ["a", "b"]
    assert doc.tables[0]["rows"] == [["1", "2"]]


# -- JSON / JSONL -------------------------------------------------------------


def test_json_object_is_structured(tmp_path):
    p = tmp_path / "obj.json"
    p.write_text('{"city": "Paris", "pop": 2}', encoding="utf-8")
    doc = load_document(p)
    assert doc.media_type == "application/json"
    assert doc.metadata["parsed"] is True
    assert "Paris" in doc.text


def test_jsonl_parses_each_line(tmp_path):
    p = tmp_path / "rows.jsonl"
    p.write_text('{"x": 1}\n\n{"x": 2}\n', encoding="utf-8")
    doc = load_document(p)
    assert doc.metadata["parsed"] is True
    # both records survive into the structured text
    assert "1" in doc.text and "2" in doc.text


def test_invalid_json_falls_back_to_raw_text(tmp_path):
    p = tmp_path / "broken.json"
    p.write_text("{not valid json,,,", encoding="utf-8")
    doc = load_document(p)
    assert doc.metadata["parsed"] is False
    assert doc.text == "{not valid json,,,"


# -- YAML ---------------------------------------------------------------------


def test_yaml_is_structured(tmp_path):
    p = tmp_path / "conf.yaml"
    p.write_text("name: vincio\nversion: 4\n", encoding="utf-8")
    doc = load_document(p)
    assert doc.media_type == "application/yaml"
    assert "vincio" in doc.text


def test_malformed_yaml_falls_back_to_raw(tmp_path):
    p = tmp_path / "bad.yml"
    # unclosed flow mapping is a YAML error → raw-text fallback branch
    p.write_text("a: [1, 2\nb: : :\n", encoding="utf-8")
    doc = load_document(p)
    assert doc.media_type == "application/yaml"
    assert doc.sections == []
    assert "a:" in doc.text


# -- code ---------------------------------------------------------------------


def test_code_file_extracts_symbols_and_imports(tmp_path):
    p = tmp_path / "mod.py"
    p.write_text("import os\n\ndef greet():\n    return 1\n", encoding="utf-8")
    doc = load_document(p)
    assert doc.media_type == "text/x-python"
    assert doc.metadata["language"] == "python"
    assert "os" in doc.metadata["imports"]
    symbol_names = [s["name"] for s in doc.metadata["symbols"]]
    assert "greet" in symbol_names


# -- email --------------------------------------------------------------------


def test_eml_extracts_headers_and_body(tmp_path):
    p = tmp_path / "msg.eml"
    p.write_bytes(
        b"From: alice@example.com\r\n"
        b"To: bob@example.com\r\n"
        b"Subject: Hello\r\n"
        b"Date: Mon, 1 Jan 2024 00:00:00 +0000\r\n"
        b"Content-Type: text/plain\r\n\r\n"
        b"This is the body.\r\n"
    )
    doc = load_document(p)
    assert doc.media_type == "message/rfc822"
    assert doc.title == "Hello"
    assert doc.metadata["from"] == "alice@example.com"
    assert "This is the body." in doc.text
    assert "Subject: Hello" in doc.text


def test_eml_without_text_body_leaves_body_empty(tmp_path):
    # A non-text Content-Type → get_body resolves to None (the 182->185 branch),
    # so the assembled text carries only the headers, no body.
    p = tmp_path / "pic.eml"
    p.write_bytes(
        b"From: x@y.com\r\nSubject: Pic\r\nContent-Type: image/png\r\n\r\nrawbytes"
    )
    doc = load_document(p)
    assert doc.title == "Pic"
    assert doc.metadata["from"] == "x@y.com"
    # body section is empty: text ends with the blank line after the headers
    assert doc.text.endswith("To: \n\n")


def test_docx_invalid_zip_wraps_in_loader_error(tmp_path):
    # Reaches the generic except branch in load_document: docx raises a
    # non-LoaderError (bad package), which is wrapped as a LoaderError.
    pytest.importorskip("docx")
    p = tmp_path / "corrupt.docx"
    p.write_bytes(b"this is not a docx zip")
    with pytest.raises(LoaderError, match="failed to load"):
        load_document(p)


# -- dispatch: not-found, unknown suffix, optional-dep wrapping ---------------


def test_load_document_missing_file_raises(tmp_path):
    with pytest.raises(LoaderError, match="file not found"):
        load_document(tmp_path / "nope.txt")


def test_unknown_suffix_best_effort_plain_text(tmp_path):
    p = tmp_path / "thing.unknownext"
    p.write_text("mystery content", encoding="utf-8")
    doc = load_document(p)
    assert doc.media_type == "text/plain"
    assert doc.text == "mystery content"


def test_load_document_sets_provenance_metadata(tmp_path):
    p = tmp_path / "prov.txt"
    p.write_text("hi", encoding="utf-8")
    doc = load_document(p, tenant_id="acme")
    assert doc.source_uri == str(p)
    assert doc.tenant_id == "acme"
    assert doc.metadata["filename"] == "prov.txt"
    assert doc.metadata["size_bytes"] == p.stat().st_size


@pytest.mark.skipif(_HAS_OPENPYXL, reason="openpyxl present; missing-dep path not reachable")
def test_xlsx_without_openpyxl_wraps_loader_error(tmp_path):
    p = tmp_path / "book.xlsx"
    p.write_bytes(b"PK\x03\x04 not a real xlsx")
    with pytest.raises(LoaderError, match="XLSX support requires openpyxl"):
        load_document(p)


def test_load_document_pdf_dispatch_without_pypdf(tmp_path):
    # Dispatch routes .pdf to load_pdf; absent pypdf surfaces as a LoaderError.
    try:
        import pypdf  # noqa: F401
    except ImportError:
        p = tmp_path / "scan.pdf"
        p.write_bytes(b"%PDF-1.4 not really")
        with pytest.raises(LoaderError, match="PDF support requires pypdf"):
            load_document(p)
    else:  # pragma: no cover - depends on env
        pytest.skip("pypdf installed")


def test_pdf_text_path_requires_pypdf():
    # pypdf is absent in the offline test env: the text path raises eagerly.
    try:
        import pypdf  # noqa: F401
    except ImportError:
        with pytest.raises(LoaderError, match="PDF support requires pypdf"):
            load_pdf("anything.pdf")
    else:  # pragma: no cover - depends on env
        pytest.skip("pypdf installed")


def test_pdf_layout_path_requires_pdfplumber():
    with pytest.raises(LoaderError, match="pdfplumber"):
        load_pdf("anything.pdf", layout=True)


def test_default_rasterizer_requires_pypdfium2():
    try:
        import pypdfium2  # noqa: F401
    except ImportError:
        with pytest.raises(LoaderError, match="requires a rasterizer"):
            _default_rasterizer("anything.pdf", 0)
    else:  # pragma: no cover - depends on env
        pytest.skip("pypdfium2 installed")


# -- DOCX (python-docx is available) ------------------------------------------


def test_docx_extracts_headings_body_and_table(tmp_path):
    docx = pytest.importorskip("docx")

    builder = docx.Document()
    builder.add_heading("Intro", level=1)
    builder.add_paragraph("Hello body text.")
    builder.add_heading("Details", level=2)
    builder.add_paragraph("Body under details.")
    table = builder.add_table(rows=2, cols=2)
    table.cell(0, 0).text = "Col1"
    table.cell(0, 1).text = "Col2"
    table.cell(1, 0).text = "v1"
    table.cell(1, 1).text = "v2"
    p = tmp_path / "report.docx"
    builder.save(str(p))

    doc = load_document(p)
    assert doc.media_type.endswith("wordprocessingml.document")
    section_titles = [s["title"] for s in doc.sections]
    assert section_titles == ["Intro", "Details"]
    # heading level parsed from the style name ("Heading 2" → 2)
    assert doc.sections[1]["level"] == 2
    # body paragraph attached to its preceding heading section
    assert "Hello body text." in doc.sections[0]["text"]
    assert doc.tables[0]["columns"] == ["Col1", "Col2"]
    assert doc.tables[0]["rows"] == [["v1", "v2"]]


def test_load_docx_direct(tmp_path):
    docx = pytest.importorskip("docx")

    builder = docx.Document()
    builder.add_paragraph("plain paragraph")
    p = tmp_path / "flat.docx"
    builder.save(str(p))
    doc = load_docx(p)
    assert doc.text == "plain paragraph"
    assert doc.sections == []  # no headings → no sections


def test_docx_skips_blank_paragraphs(tmp_path):
    docx = pytest.importorskip("docx")

    builder = docx.Document()
    builder.add_paragraph("real line")
    builder.add_paragraph("")  # blank → skipped by the `if not text: continue` guard
    builder.add_paragraph("   ")  # whitespace-only → also skipped
    builder.add_paragraph("second line")
    p = tmp_path / "gappy.docx"
    builder.save(str(p))
    doc = load_docx(p)
    # only the two non-blank paragraphs survive
    assert doc.text == "real line\nsecond line"


# -- audio / media ------------------------------------------------------------


def test_load_media_builds_timestamped_sections(tmp_path):
    p = tmp_path / "talk.wav"
    p.write_bytes(b"RIFFfake")
    doc = load_media(p, transcriber=MockTranscriber(), tenant_id="t1")
    assert doc.media_type == "audio/transcript"
    assert doc.tenant_id == "t1"
    assert doc.metadata["segment_count"] == 2
    assert doc.metadata["language"] == "en"
    assert doc.metadata["extractor"] == "transcript"
    # timestamp prefix is rendered into each section's text
    assert doc.sections[0]["text"].startswith("[")
    assert doc.sections[0]["speaker"] is None


def test_load_media_diarized_carries_speaker(tmp_path):
    p = tmp_path / "talk.wav"
    p.write_bytes(b"x")
    doc = load_media(
        p, transcriber=MockTranscriber(script=["one", "two"], diarize=True)
    )
    assert doc.sections[0]["speaker"] == "Speaker 1"
    assert doc.sections[1]["speaker"] == "Speaker 2"
    assert "Speaker 1:" in doc.sections[0]["text"]


def test_load_media_missing_file_raises(tmp_path):
    with pytest.raises(LoaderError, match="file not found"):
        load_media(tmp_path / "absent.wav", transcriber=MockTranscriber())


# -- video --------------------------------------------------------------------


def test_load_video_builds_segments_with_frames(tmp_path):
    p = tmp_path / "clip.mp4"
    p.write_bytes(b"\x00\x00\x00fake")
    doc = load_video(p, analyzer=MockVideoAnalyzer(), tenant_id="v")
    assert doc.media_type == "video/analysis"
    assert doc.tenant_id == "v"
    assert doc.metadata["segment_count"] == 3
    assert doc.metadata["fps"] == 1.0
    assert doc.sections[0]["frame_count"] == 1
    assert doc.sections[0]["extractor"] == "video"
    # second segment starts right after the first (5s segments)
    assert doc.sections[1]["start"] == 5.0


def test_load_video_missing_file_raises(tmp_path):
    with pytest.raises(LoaderError, match="file not found"):
        load_video(tmp_path / "absent.mp4", analyzer=MockVideoAnalyzer())


# -- figure_evidence ----------------------------------------------------------


class _StubImageAnalyzer:
    """Minimal concrete analyzer: returns fixed region-tagged observations."""

    def observe(self, image_path):  # noqa: D401 - sync; loader awaits-or-passes
        return [
            ImageObservation(region="top", observation="a bar chart", confidence=0.7),
            ImageObservation(region="bottom", observation="a caption", confidence=0.6),
        ]


class _StubOCR:
    """Minimal concrete OCR engine returning the text we configure."""

    def __init__(self, text):
        self._text = text

    def extract_text(self, image_path):
        return self._text


def _doc_with_figure() -> Document:
    return Document(
        text="paper",
        title="paper",
        metadata={"figures": [{"bbox": [0, 0, 10, 10], "page": 3}]},
    )


def test_figure_evidence_with_analyzer():
    doc = _doc_with_figure()
    items = figure_evidence(doc, crops={0: "/crop0.png"}, analyzer=_StubImageAnalyzer())
    assert len(items) == 2
    first = items[0]
    assert first.id == f"{doc.id}:FIG0:R1"
    assert first.text == "[figure 0] a bar chart"
    assert first.page == 3
    assert first.authority == 0.7
    assert first.metadata["bbox"] == [0, 0, 10, 10]
    assert first.metadata["region"] == "top"


def test_figure_evidence_with_ocr():
    doc = _doc_with_figure()
    items = figure_evidence(doc, crops={0: "/crop0.png"}, ocr_engine=_StubOCR("Sales: 42"))
    assert len(items) == 1
    assert items[0].text == "[figure 0] Sales: 42"
    assert items[0].metadata["extractor"] == "ocr"
    assert items[0].page == 3


def test_figure_evidence_ocr_blank_yields_nothing():
    doc = _doc_with_figure()
    items = figure_evidence(doc, crops={0: "/crop0.png"}, ocr_engine=_StubOCR("   "))
    assert items == []


def test_figure_evidence_out_of_range_index_uses_empty_figure():
    doc = _doc_with_figure()
    # index 9 has no matching figure → bbox/page resolve to None
    items = figure_evidence(doc, crops={9: "/x.png"}, ocr_engine=_StubOCR("text"))
    assert items[0].page is None
    assert items[0].metadata["bbox"] is None


def test_figure_evidence_no_engine_returns_empty():
    doc = _doc_with_figure()
    assert figure_evidence(doc, crops={0: "/x.png"}) == []


# -- load_directory -----------------------------------------------------------


def test_load_directory_walks_and_builds_repo_summary(tmp_path):
    repo = tmp_path / "repo"
    (repo / "pkg").mkdir(parents=True)
    (repo / "pkg" / "a.py").write_text("import os\nimport sys\n", encoding="utf-8")
    # a code file with NO imports: its (empty) import list is skipped in the
    # summary loop's `if imports:` branch (the 651->650 false edge).
    (repo / "pkg" / "b.py").write_text("def f():\n    return 1\n", encoding="utf-8")
    (repo / "readme.md").write_text("# Repo\n\nhello\n", encoding="utf-8")
    # ignored dir + unsupported suffix are skipped
    (repo / "__pycache__").mkdir()
    (repo / "__pycache__" / "junk.py").write_text("x=1\n", encoding="utf-8")
    (repo / "image.bin").write_bytes(b"\x00\x01")

    docs = load_directory(repo, tenant_id="org")
    titles = {d.title for d in docs}
    assert "a.py" in titles
    assert "readme" in titles
    assert "junk.py" not in titles  # ignored __pycache__
    # repo summary appended because code files were found
    summary = next(d for d in docs if d.metadata.get("kind") == "repo_summary")
    assert "import graph" in summary.text.lower()
    assert "os" in summary.text and "sys" in summary.text
    # the import-free module is not listed in the graph
    assert "b.py ->" not in summary.text
    assert summary.tenant_id == "org"
    # relative_path recorded on each real document
    code_doc = next(d for d in docs if d.title == "a.py")
    assert code_doc.metadata["relative_path"] == str(Path("pkg") / "a.py")


def test_load_directory_no_code_has_no_summary(tmp_path):
    repo = tmp_path / "prose"
    repo.mkdir()
    (repo / "one.txt").write_text("alpha", encoding="utf-8")
    (repo / "two.txt").write_text("beta", encoding="utf-8")
    docs = load_directory(repo)
    assert len(docs) == 2
    assert all(d.metadata.get("kind") != "repo_summary" for d in docs)


def test_load_directory_respects_extension_filter(tmp_path):
    repo = tmp_path / "mixed"
    repo.mkdir()
    (repo / "keep.md").write_text("# keep", encoding="utf-8")
    (repo / "drop.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    docs = load_directory(repo, extensions={".md"})
    assert {d.title for d in docs} == {"keep"}


def test_load_directory_max_files_caps_results(tmp_path):
    repo = tmp_path / "many"
    repo.mkdir()
    for i in range(5):
        (repo / f"f{i}.txt").write_text(str(i), encoding="utf-8")
    docs = load_directory(repo, max_files=2)
    assert len(docs) == 2


def test_load_directory_custom_ignore_dir(tmp_path):
    repo = tmp_path / "r"
    (repo / "vendor").mkdir(parents=True)
    (repo / "vendor" / "lib.txt").write_text("vendored", encoding="utf-8")
    (repo / "main.txt").write_text("main", encoding="utf-8")
    docs = load_directory(repo, ignore_dirs={"vendor"})
    assert {d.title for d in docs} == {"main"}


def test_load_directory_missing_dir_raises(tmp_path):
    with pytest.raises(LoaderError, match="directory not found"):
        load_directory(tmp_path / "ghost")


def test_load_directory_skips_unreadable_optional_dep_file(tmp_path):
    # A .xlsx that needs openpyxl (absent) is skipped, not fatal, in bulk load.
    repo = tmp_path / "bulk"
    repo.mkdir()
    (repo / "ok.txt").write_text("fine", encoding="utf-8")
    if not _HAS_OPENPYXL:
        (repo / "bad.xlsx").write_bytes(b"not a workbook")
    docs = load_directory(repo)
    titles = {d.title for d in docs}
    assert "ok" in titles
    assert "bad" not in titles


# -- supported_extensions -----------------------------------------------------


def test_supported_extensions_includes_builtins():
    exts = supported_extensions()
    for ext in (".txt", ".md", ".csv", ".json", ".yaml", ".html", ".py", ".pdf", ".docx", ".xlsx"):
        assert ext in exts


def test_document_error_importable():
    # DocumentError participates in the OCR re-raise contract within load_pdf.
    assert issubclass(DocumentError, Exception)
